from __future__ import annotations

import io
import math
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import pytesseract


def _ensure_tesseract_cmd() -> None:
    candidates = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return


def _resolve_ocr_lang(requested_lang: str) -> str:
    available = set(pytesseract.get_languages(config=""))
    if not available:
        raise RuntimeError("Tesseract terpasang tetapi tidak menemukan data bahasa OCR.")

    requested_parts = [p.strip() for p in requested_lang.split("+") if p.strip()]
    if not requested_parts:
        requested_parts = ["eng"]

    selected_parts = [p for p in requested_parts if p in available]
    if selected_parts:
        return "+".join(selected_parts)

    if "eng" in available:
        return "eng"
    return next(iter(available))


def _extract_ocr_text(image: PILImage.Image, ocr_lang: str) -> str:
    text = pytesseract.image_to_string(image, lang=ocr_lang)
    return text.strip()


def _set_columns_for_width(ws, image_width_px: int) -> None:
    column_width = 12
    approx_px_per_column = int(column_width * 7)
    needed_cols = max(1, math.ceil(image_width_px / approx_px_per_column))

    for col in range(1, min(needed_cols, 80) + 1):
        ws.column_dimensions[get_column_letter(col)].width = column_width


def pdf_to_excel_preserve_layout(
    pdf_bytes: bytes,
    output_xlsx: Path,
    dpi: int = 200,
    with_ocr: bool = False,
    ocr_lang: str = "ind+eng",
) -> None:
    if dpi < 72:
        raise ValueError("dpi minimal 72")

    if with_ocr:
        _ensure_tesseract_cmd()
        # Fail fast if Tesseract binary is not available in PATH.
        try:
            _ = pytesseract.get_tesseract_version()
        except Exception as exc:
            raise RuntimeError(
                "OCR aktif tapi Tesseract OCR belum terpasang atau belum masuk PATH."
            ) from exc
        ocr_lang = _resolve_ocr_lang(ocr_lang)

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    wb = Workbook()
    sheet = wb.active
    sheet.title = "PDF"
    sheet.sheet_view.showGridLines = False
    image_streams = []

    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    current_row = 1
    max_width = 0

    for page_index, page in enumerate(doc, start=1):
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        png_bytes = pix.tobytes("png")
        image = PILImage.open(io.BytesIO(png_bytes))
        png_stream = io.BytesIO(png_bytes)
        image_streams.append(png_stream)
        excel_image = XLImage(png_stream)
        sheet.add_image(excel_image, f"A{current_row}")

        max_width = max(max_width, pix.width)
        approx_px_per_row = 20
        page_rows = max(1, math.ceil(pix.height / approx_px_per_row))
        for row in range(current_row, current_row + page_rows + 1):
            sheet.row_dimensions[row].height = 15
        current_row += page_rows + 2

        if with_ocr:
            ocr_header = sheet.cell(row=current_row, column=1, value=f"OCR Page {page_index}")
            ocr_header.font = Font(bold=True)
            current_row += 1

            ocr_text = _extract_ocr_text(image, ocr_lang) or "(tidak ada teks terdeteksi)"
            ocr_cell = sheet.cell(row=current_row, column=1, value=ocr_text)
            ocr_cell.alignment = Alignment(wrap_text=True, vertical="top")

            line_count = max(3, min(250, ocr_text.count("\n") + 1))
            for row in range(current_row, current_row + line_count):
                sheet.row_dimensions[row].height = 15
            current_row += line_count + 2

    _set_columns_for_width(sheet, max_width or 1200)

    wb.save(output_xlsx)
